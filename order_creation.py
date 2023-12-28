# -*- coding: utf-8 -*-
import pandas as pd
import pymongo
################################################################################
## Form generated from reading UI file 'order_creation.ui'
##
## Created by: Qt User Interface Compiler version 6.6.1
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QHeaderView, QLabel, QPushButton,
                               QSizePolicy, QTableWidget, QTableWidgetItem, QWidget, QDialog, QMessageBox)
from sqlalchemy import create_engine, text

import docs_browser
import order_add_position
import parameters

import pandas as pd
import os
import shutil

from time import time, sleep

from sqlalchemy import text, create_engine
from openpyxl import load_workbook

from openpyxl.styles import Alignment
from parameters import DB_PATH, EXCEL_TEMPLATE_PATH, MAIN_REPORT_PAGE, EXCEL_HEADER_ROWS, SAVE_DIR, WRAP_COLUMNS, NAMES_TXT_PATH



class Ui_Form(object):
    def refresh_order_entries_table(self):
        parameters.refresh_order_entries_table(self)

    def create_report(self):
        print("SHIT STARTED")
        self.vet_db_connection = create_engine(f'sqlite:///{DB_PATH}').connect()
        supplier_text = parameters.get_supplier_for_order(self.order_num)
        customer_text = parameters.get_customer_for_order(self.order_num)

        report_entries = pd.read_sql(text(
            f'''SELECT pk, name, count, unit, price, sum FROM order_entry WHERE related_to_order={self.order_num}'''#HYETA MOZHET NE SRABOTAT
            # f'''SELECT household.owner, city.name, household.address, report_entries.specie, report_entries.count, report_entries.data_from_administration, report_entries.prevous_count, report_entries.is_conditions_good FROM report_entries
            #     --присоединяем данные о хозяйстве (владелец и его адрес в рамках города)
            #     INNER JOIN household
            #     ON household.pk = report_entries.household
            #
            #     --присоединяем данные о городе (для хозяйства, уточняем адрес)
            #     INNER JOIN city
            #     ON city.pk = household.belongs_to_city
            #     WHERE city.pk = {city}
            #
            # '''
        ), self.vet_db_connection)

        #self.sum = report_entries["sum"].sum()
        #self.total_count = report_entries["count"].sum()

        print(report_entries)
        # ----------------------------------------------------------------------------------дефолтные прелбразования данныхз ради некривого форматирования--------------------------------------------------------------------------------------------------------------------------
        # report_entries.insert(0, 'position', report_entries['owner'].astype('category').cat.codes + 1)
        # report_entries = report_entries.sort_values('owner')
        # report_entries.loc[
        #     report_entries[['owner']].duplicated(keep='first'), ['owner', 'address', 'name', 'position']] = ''
        # report_entries['address'] = report_entries.apply(
        #     lambda x: f"{x['name']}, {x['address']}" if x['address'] else "", axis=1)
        # report_entries = report_entries.drop(columns='name')
        # print(report_entries)

        # ----------------------------------------------------------------------------------создаем диреткории если нет и копируем шаблон, затем заполняем его--------------------------------------------------------------------------------------------------------------------------
        os.makedirs(SAVE_DIR, exist_ok=True)
        filename = os.path.join(SAVE_DIR, f'Отчёт{int(time())}.xlsx')
        shutil.copy(EXCEL_TEMPLATE_PATH, filename)
        current_report = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # заполнение данными
        # with open(NAMES_TXT_PATH, 'r', encoding='utf-8') as names_file:
        #     names = [line.strip() for line in names_file.readlines()]
        names = {
            'ORDERN': self.order_num, 'ORDERD': self.date, 'SUPPLIERPH': parameters.get_supplier_for_order(self.order_num), 'CUSTOMERPH': parameters.get_customer_for_order(self.order_num),
            'COUNT': str(report_entries["count"].sum()),'SUM11TOTAL': str(report_entries["sum"].sum()),'SUMTOTALNDS': str(report_entries["sum"].sum()*0.18),
            'SUMSUMTOTALPLUS': str(report_entries["sum"].sum()*1.18)
                 }

        page = current_report.sheets[MAIN_REPORT_PAGE]
        message = QMessageBox()
        message.setText("loh")
        message.show()
        sleep(1.5)

        def fill_placeholders(names, cell):
            if isinstance(cell.value, str):
                for placeholder, name in names.items():
                    cell.value = cell.value.replace(placeholder, name)

        for row in range(1, page.max_row + 1):
            for col in range(1, page.max_column + 1):
                fill_placeholders(names, page.cell(row, col))

        page.insert_rows(EXCEL_HEADER_ROWS + 1, amount=report_entries.shape[0])
        for row in range(report_entries.shape[0]):
            for col, col_name in enumerate(report_entries.columns):
                cell = page.cell(row + EXCEL_HEADER_ROWS + 1, col + 1)
                cell.value = report_entries.iloc[row, col]
                # перенос по словам
                if col_name in WRAP_COLUMNS:
                    cell.alignment = cell.alignment.copy(wrapText=True)

        current_report.close()

    def change_order(self, order_num):
        #print("vot tut nado menyat postavshikai pokupatela v shapke")
        self.order_num = order_num
        self.label_5.setText(str(self.order_num))
        self.refresh_order_entries_table()
        self.label_4.setText(parameters.get_supplier_for_order(order_num))
        self.label_7.setText(parameters.get_customer_for_order(order_num))

    def __init__(self):
        #self.refresh_order_entries_table = parameters.refresh_order_entries_table()
        self.order_num = parameters.just_give_doc_number()#pk+1 dayot nomer dokumenta dalya interfeisa, iz nego mozhno poluchit pk documenta otnyav 1

    def create_new_order(self):
        parameters.create_new_order()
        #new_order_pk = new_order_number-1
        self.order_num=parameters.just_give_doc_number()
        self.label_5.setText(str(self.order_num))
        self.refresh_order_entries_table()
        client = pymongo.MongoClient("localhost", 27017)  # mongo penis
        db = client["kursach_ais"]
        pk_collection = db["orderj"]
        pk_collection.insert_many([
            {"pk": self.order_num, "kostyl": 'ass', "date": self.label_6, "supplier": self.supplier, "customer": self.customer},
        ]);

    def open_docs_browser(self):#add order entry window
        self.window = QDialog()
        self.ui = docs_browser.Ui_widget()
        self.ui.transfer_form_pointer(self)
        self.ui.setupUi(self.window)
        self.window.show()

    def open_order_add_position(self):#add order entry window
        self.window = QDialog()
        self.ui = order_add_position.Ui_Form()
        self.ui.transfer_data(self.order_num, self)#vozmozhno oshibka i nado otnyat 1, ya v ujase pochemy ono rabotaet
        self.ui.setupUi(self.window)
        self.window.show()
    def transfer_data(self, customer, supplier, date):
        self.customer=customer
        self.supplier=supplier
        self.date=date



    def setupUi(self, Form):
        if not Form.objectName():
            Form.setObjectName(u"Form")
        Form.resize(1017, 695)
        self.label = QLabel(Form)
        self.label.setObjectName(u"label")
        self.label.setGeometry(QRect(10, 10, 741, 61))
        font = QFont()
        font.setPointSize(20)
        font.setBold(True)
        self.label.setFont(font)
        self.label_2 = QLabel(Form)
        self.label_2.setObjectName(u"label_2")
        self.label_2.setGeometry(QRect(20, 70, 91, 61))
        font1 = QFont()
        font1.setPointSize(11)
        font1.setBold(False)
        self.label_2.setFont(font1)
        self.label_3 = QLabel(Form)
        self.label_3.setObjectName(u"label_3")
        self.label_3.setGeometry(QRect(20, 110, 91, 61))
        self.label_3.setFont(font1)
        self.label_5 = QLabel(Form) #order number
        self.label_5.setObjectName(u"label_5")
        self.label_5.setGeometry(QRect(290, 10, 81, 61))
        self.label_5.setFont(font)
        self.label_5.setText(self.order_num)

        self.label_6 = QLabel(Form)#order date
        self.label_6.setObjectName(u"label_6")
        self.label_6.setGeometry(QRect(420, 10, 201, 61))
        self.label_6.setFont(font)
        self.label_6.setText(self.date)


        self.label_4 = QLabel(Form)#supplier
        self.label_4.setObjectName(u"label_4")
        self.label_4.setGeometry(QRect(110, 90, 691, 21))
        self.label_4.setFont(font1)
        self.label_4.setText(self.supplier)

        self.label_7 = QLabel(Form)#customer
        self.label_7.setObjectName(u"label_7")
        self.label_7.setGeometry(QRect(110, 130, 781, 21))
        self.label_7.setFont(font1)
        self.label_7.setText(self.customer)

        self.positionsTableWidget = QTableWidget(Form)
        if (self.positionsTableWidget.columnCount() < 5):
            self.positionsTableWidget.setColumnCount(5)
        __qtablewidgetitem = QTableWidgetItem()
        self.positionsTableWidget.setHorizontalHeaderItem(0, __qtablewidgetitem)
        __qtablewidgetitem1 = QTableWidgetItem()
        self.positionsTableWidget.setHorizontalHeaderItem(1, __qtablewidgetitem1)
        __qtablewidgetitem2 = QTableWidgetItem()
        self.positionsTableWidget.setHorizontalHeaderItem(2, __qtablewidgetitem2)
        __qtablewidgetitem3 = QTableWidgetItem()
        self.positionsTableWidget.setHorizontalHeaderItem(3, __qtablewidgetitem3)
        __qtablewidgetitem4 = QTableWidgetItem()
        self.positionsTableWidget.setHorizontalHeaderItem(4, __qtablewidgetitem4)
        self.positionsTableWidget.setObjectName(u"positionsTableWidget")
        self.positionsTableWidget.setGeometry(QRect(20, 190, 721, 371))
        self.addPushButton = QPushButton(Form, clicked = lambda: self.open_order_add_position())
        self.addPushButton.setObjectName(u"addPushButton")
        self.addPushButton.setGeometry(QRect(20, 580, 141, 81))
        self.deletePushButton = QPushButton(Form)
        self.deletePushButton.setObjectName(u"deletePushButton")
        self.deletePushButton.setGeometry(QRect(170, 580, 141, 81))
        self.createOrderPushButton = QPushButton(Form, clicked = lambda:self.create_report())
        self.createOrderPushButton.setObjectName(u"createOrderPushButton")
        self.createOrderPushButton.setGeometry(QRect(600, 580, 141, 81))
        self.newOrderPushButton = QPushButton(Form, clicked = lambda:self.create_new_order())#new order idk
        self.newOrderPushButton.setObjectName(u"createOrderPushButton_2")
        self.newOrderPushButton.setGeometry(QRect(870, 20, 141, 81))

        self.openDocsBrowserPushButton = QPushButton(Form, clicked=lambda: self.open_docs_browser())  # new order idk
        self.openDocsBrowserPushButton.setObjectName(u"createOrderPushButton_2")
        self.openDocsBrowserPushButton.setGeometry(QRect(800, 160, 141, 81))

        self.refresh_order_entries_table()

        self.retranslateUi(Form)

        QMetaObject.connectSlotsByName(Form)
    # setupUi

    def retranslateUi(self, Form):
        Form.setWindowTitle(QCoreApplication.translate("Form", u"Form", None))
        self.label.setText(QCoreApplication.translate("Form", u"\u0417\u0430\u043a\u0430\u0437 \u043f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a\u0443 \u2116             \u043e\u0442 ", None))
        self.label_2.setText(QCoreApplication.translate("Form", u"\u041f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a:", None))
        self.label_3.setText(QCoreApplication.translate("Form", u"\u041f\u043e\u043a\u0443\u043f\u0430\u0442\u0435\u043b\u044c:", None))
        # self.label_5.setText(QCoreApplication.translate("Form", u"N", None))
        # self.label_6.setText(QCoreApplication.translate("Form", u"DATE", None))
        # self.label_4.setText(QCoreApplication.translate("Form", u"supplier", None))
        # self.label_7.setText(QCoreApplication.translate("Form", u"customer", None))
        ___qtablewidgetitem = self.positionsTableWidget.horizontalHeaderItem(0)
        ___qtablewidgetitem.setText(QCoreApplication.translate("Form", u"\u043d\u043e\u043c", None));
        ___qtablewidgetitem1 = self.positionsTableWidget.horizontalHeaderItem(1)
        ___qtablewidgetitem1.setText(QCoreApplication.translate("Form", u"\u0418\u043c\u044f", None));
        ___qtablewidgetitem2 = self.positionsTableWidget.horizontalHeaderItem(2)
        ___qtablewidgetitem2.setText(QCoreApplication.translate("Form", u"\u0446\u0435\u043d\u0430", None));
        ___qtablewidgetitem3 = self.positionsTableWidget.horizontalHeaderItem(3)
        ___qtablewidgetitem3.setText(QCoreApplication.translate("Form", u"\u043a\u043e\u043b-\u0432\u043e", None));
        ___qtablewidgetitem4 = self.positionsTableWidget.horizontalHeaderItem(4)
        ___qtablewidgetitem4.setText(QCoreApplication.translate("Form", u"\u0441\u0443\u043c\u043c\u0430", None));
        self.addPushButton.setText(QCoreApplication.translate("Form", u"\u0414\u043e\u0431\u0430\u0432\u0438\u0442\u044c ", None))
        self.deletePushButton.setText(QCoreApplication.translate("Form", u"\u0423\u0434\u0430\u043b\u0438\u0442\u044c", None))
        self.createOrderPushButton.setText(QCoreApplication.translate("Form", u"\u0412\u044b\u043f\u0443\u0441\u0442\u0438\u0442\u044c", None))
        self.newOrderPushButton.setText("новый")
        self.openDocsBrowserPushButton.setText("МЕНЕДЖЕР")
    # retranslateUi

